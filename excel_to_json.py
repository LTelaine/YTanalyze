"""
Excel → JSON 轉換腳本
======================
讀取 醍醐WAY_YT_數據_模板_v3.xlsx，輸出 data.json 供 Dashboard 使用。

使用方式：
  python excel_to_json.py
  python excel_to_json.py --input 其他檔案.xlsx --output 其他路徑/data.json
"""

import argparse
import json
import os
import re
from openpyxl import load_workbook

DEFAULT_INPUT = "醍醐WAY_YT_數據_模板_v3.xlsx"
DEFAULT_OUTPUT = os.path.join("YTanalyze", "public", "data.json")

# Excel 欄位對應 (1-based column index)
# cols 1-18: API 自動欄位
# cols 19-33: 手動填寫（集數、AB test、選題）
# cols 34-39: 影片類型、來賓、選題大類、計算欄位
# 注意：目前 Excel 資料 cols 34-39 偏移一欄（col 34 空白，實際資料從 col 35 開始）
COL = {
    "id": 1, "show": 2, "title": 3, "date": 4,
    "views": 6, "avgWatch": 9, "watchMin": 10,
    "subs": 11, "likes": 12, "comments": 13, "shares": 14,
    "age": 15, "female": 16, "male": 17, "traffic": 18,
    "ep": 19,
    "copyA": 20, "ctrA": 21, "copyB": 22, "ctrB": 23,
    "winner": 24, "frameA": 25, "frameB": 26,
    "testVar": 27, "angleA": 28, "angleB": 29,
    "conclusion": 30, "suggestion": 31,
    "topicAB": 32,
}

# cols 34-39 有偏移，實際資料位置
COL_SHIFTED = {
    "type": 35,
    "guest": 36,
    "topic": 37,
    "interactRate": 38,
    "subsRate": 39,
}

VIDEO_TYPES = {"完整集", "Shorts", "Podcast", "精華版"}


def cell(ws, row, col):
    v = ws.cell(row, col).value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v in ("", "N/A", "（需 Reach Report）"):
            return None
    return v


def parse_watch_time(val):
    """'13:37' → '13:37', None → '0:00'"""
    if not val:
        return "0:00"
    return str(val)


def detect_shift(ws, max_row):
    """偵測 cols 34-39 是否有偏移"""
    for r in range(3, min(max_row + 1, 20)):
        v34 = ws.cell(r, 34).value
        v35 = ws.cell(r, 35).value
        if v35 and str(v35).strip() in VIDEO_TYPES:
            if not v34 or str(v34).strip() not in VIDEO_TYPES:
                return True
    return False


def extract_guest_name(raw):
    """'蕭敦仁醫師 EP3' → '蕭敦仁', '預防科科長 莊明雄 EP16' → '莊明雄'"""
    if not raw:
        return ""
    s = str(raw).strip()
    s = re.sub(r"\s*EP\d+.*$", "", s)
    s = re.sub(r"醫師|律師|教授|老師|科長|主任|博士|教練", "", s)
    parts = s.strip().split()
    if len(parts) > 1:
        return parts[-1]
    return s.strip()


def convert(input_path, output_path):
    print(f"📖 讀取 {input_path}")
    wb = load_workbook(input_path, data_only=True)
    ws = wb["影片數據（API自動）"]

    shifted = detect_shift(ws, ws.max_row)
    if shifted:
        print("   偵測到 cols 34-39 資料偏移，自動校正")
        type_col, guest_col, topic_col = 35, 36, 37
    else:
        type_col, guest_col, topic_col = 34, 35, 36

    videos = []
    ab_tests = []

    for r in range(3, ws.max_row + 1):
        vid = cell(ws, r, COL["id"])
        if not vid:
            continue

        views = cell(ws, r, COL["views"]) or 0
        likes = cell(ws, r, COL["likes"]) or 0
        comments_val = cell(ws, r, COL["comments"]) or 0
        shares = cell(ws, r, COL["shares"]) or 0
        subs = cell(ws, r, COL["subs"]) or 0
        female = cell(ws, r, COL["female"]) or 0
        male = cell(ws, r, COL["male"]) or 0

        video_type = str(cell(ws, r, type_col) or "完整集").strip()
        guest_raw = cell(ws, r, guest_col)
        topic = str(cell(ws, r, topic_col) or "未分類").strip()
        guest = extract_guest_name(guest_raw) if video_type == "完整集" else ""

        video = {
            "id": str(vid),
            "show": str(cell(ws, r, COL["show"]) or ""),
            "title": str(cell(ws, r, COL["title"]) or ""),
            "date": str(cell(ws, r, COL["date"]) or ""),
            "views": int(views),
            "watchMin": int(cell(ws, r, COL["watchMin"]) or 0),
            "subs": int(subs),
            "likes": int(likes),
            "comments": int(comments_val),
            "shares": int(shares),
            "avgWatch": parse_watch_time(cell(ws, r, COL["avgWatch"])),
            "age": str(cell(ws, r, COL["age"]) or ""),
            "female": float(female),
            "male": float(male),
            "traffic": str(cell(ws, r, COL["traffic"]) or "N/A"),
            "type": video_type,
            "guest": guest,
            "topic": topic,
            "ep": str(cell(ws, r, COL["ep"]) or ""),
        }
        videos.append(video)

        # AB test: 有 copyA 和 copyB 才算有效
        copyA = cell(ws, r, COL["copyA"])
        copyB = cell(ws, r, COL["copyB"])
        if copyA and copyB:
            ab = {
                "ep": video["ep"],
                "show": video["show"],
                "topic": cell(ws, r, COL["topicAB"]) or topic,
                "copyA": str(copyA),
                "copyB": str(copyB),
                "ctrA": float(cell(ws, r, COL["ctrA"]) or 0),
                "ctrB": float(cell(ws, r, COL["ctrB"]) or 0),
                "winner": str(cell(ws, r, COL["winner"]) or ""),
                "frameA": str(cell(ws, r, COL["frameA"]) or ""),
                "frameB": str(cell(ws, r, COL["frameB"]) or ""),
                "testVar": str(cell(ws, r, COL["testVar"]) or ""),
                "angleA": str(cell(ws, r, COL["angleA"]) or ""),
                "angleB": str(cell(ws, r, COL["angleB"]) or ""),
                "conclusion": str(cell(ws, r, COL["conclusion"]) or ""),
                "suggestion": str(cell(ws, r, COL["suggestion"]) or ""),
            }
            ab_tests.append(ab)

    data = {"videos": videos, "abTests": ab_tests}

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    full_count = sum(1 for v in videos if v["type"] == "完整集")
    print(f"✅ 輸出 {output_path}")
    print(f"   影片：{len(videos)} 筆（完整集 {full_count}）")
    print(f"   AB test：{len(ab_tests)} 筆")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Excel → JSON for Dashboard")
    parser.add_argument("--input", default=DEFAULT_INPUT, help="Excel 檔案路徑")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="JSON 輸出路徑")
    args = parser.parse_args()
    convert(args.input, args.output)
